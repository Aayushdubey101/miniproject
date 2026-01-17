import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import msoffcrypto
from io import BytesIO
from win32com.client.gencache import EnsureDispatch
from win32com.client import Dispatch
import pythoncom
import plotly.express as px
import plotly.graph_objects as go
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import zipfile
import tempfile
import os
from copy import copy

# Page configuration
st.set_page_config(page_title="Excel Manipulation Tool", layout="wide", page_icon="📊")

# Initialize session state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'workbook' not in st.session_state:
    st.session_state.workbook = None
if 'file_path' not in st.session_state:
    st.session_state.file_path = None
if 'df_dict' not in st.session_state:
    st.session_state.df_dict = {}

# ==================== HELPER FUNCTIONS ====================

@st.cache_data
def load_excel_with_password(file_bytes, password=None):
    """Load Excel file with optional password protection"""
    try:
        if password:
            decrypted = BytesIO()
            file_io = BytesIO(file_bytes)
            office_file = msoffcrypto.OfficeFile(file_io)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return decrypted
        else:
            return BytesIO(file_bytes)
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None

def get_all_sheets(file_io):
    """Get all sheet names from Excel file"""
    try:
        wb = load_workbook(file_io, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        st.error(f"Error reading sheets: {str(e)}")
        return []

def load_sheet_data(file_io, sheet_name=None):
    """Load data from specific sheet"""
    try:
        df = pd.read_excel(file_io, sheet_name=sheet_name, engine='openpyxl')
        return df
    except Exception as e:
        st.error(f"Error loading sheet data: {str(e)}")
        return None

def create_download_link(wb, filename):
    """Create download button for workbook"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def validate_sheet_name(name, existing_sheets):
    """Validate sheet name"""
    if not name or name.strip() == "":
        return False, "Sheet name cannot be empty"
    if name in existing_sheets:
        return False, "Sheet name already exists"
    if len(name) > 31:
        return False, "Sheet name must be 31 characters or less"
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    if any(char in name for char in invalid_chars):
        return False, f"Sheet name cannot contain: {', '.join(invalid_chars)}"
    return True, "Valid"

# ==================== BASIC OPERATIONS FUNCTIONS ====================

def create_new_excel(name):
    """Create a new Excel file"""
    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="📥 Download New Excel File",
            data=output.getvalue(),
            file_name=f"{name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success(f"Excel file '{name}.xlsx' created successfully!")
    except Exception as e:
        st.error(f"Error creating file: {str(e)}")

def modify_excel_cell(wb, sheet_name, address, value):
    """Modify a specific cell in Excel"""
    try:
        sheet = wb[sheet_name]
        sheet[address] = value
        st.success(f"Value '{value}' has been written to cell '{address}' in sheet '{sheet_name}'.")
        return wb
    except Exception as e:
        st.error(f"Error modifying cell: {str(e)}")
        return wb

def set_password_excel(file_path, password):
    """Set password for Excel file (Windows only)"""
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()
        
        xl_file = EnsureDispatch("Excel.Application")
        wb = xl_file.Workbooks.Open(file_path)
        xl_file.DisplayAlerts = False
        wb.Visible = False
        wb.SaveAs(file_path, Password=password)
        wb.Close()
        xl_file.Quit()
        
        # Uninitialize COM
        pythoncom.CoUninitialize()
        
        st.success("Password has been set successfully.")
    except Exception as e:
        pythoncom.CoUninitialize()
        st.error(f"Error setting password: {str(e)}")

def remove_password_excel(file_path, password):
    """Remove password from Excel file (Windows only)"""
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()
        
        excel_app = Dispatch("Excel.Application")
        workbook = excel_app.Workbooks.Open(file_path, False, True, None, password)
        for sheet in workbook.Worksheets:
            if sheet.ProtectContents:
                sheet.Unprotect(password)
        excel_app.DisplayAlerts = False
        
        # Save without password
        output_path = file_path.replace('.xlsx', '_unprotected.xlsx')
        workbook.SaveAs(output_path, FileFormat=51, Password="")
        workbook.Close(SaveChanges=True)
        excel_app.Quit()
        
        # Uninitialize COM
        pythoncom.CoUninitialize()
        
        # Read and provide download
        with open(output_path, 'rb') as f:
            st.download_button(
                label="📥 Download Unprotected File",
                data=f.read(),
                file_name=os.path.basename(output_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.success("Password removed successfully!")
        
        # Cleanup
        os.remove(output_path)
    except Exception as e:
        pythoncom.CoUninitialize()
        st.error(f"Error removing password: {str(e)}")

# ==================== DATA ANALYSIS FUNCTIONS ====================

def create_chart(df, chart_type, x_col, y_col, title="Chart"):
    """Create interactive charts using Plotly"""
    try:
        if chart_type == "Bar Chart":
            fig = px.bar(df, x=x_col, y=y_col, title=title)
        elif chart_type == "Line Chart":
            fig = px.line(df, x=x_col, y=y_col, title=title)
        elif chart_type == "Pie Chart":
            fig = px.pie(df, names=x_col, values=y_col, title=title)
        elif chart_type == "Scatter Plot":
            fig = px.scatter(df, x=x_col, y=y_col, title=title)
        else:
            return None
        
        fig.update_layout(template="plotly_white")
        return fig
    except Exception as e:
        st.error(f"Error creating chart: {str(e)}")
        return None

def calculate_statistics(df, columns):
    """Calculate statistics for selected columns"""
    try:
        stats_dict = {}
        for col in columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                stats_dict[col] = {
                    'Mean': df[col].mean(),
                    'Median': df[col].median(),
                    'Mode': df[col].mode()[0] if not df[col].mode().empty else None,
                    'Sum': df[col].sum(),
                    'Count': df[col].count(),
                    'Min': df[col].min(),
                    'Max': df[col].max(),
                    'Std Dev': df[col].std()
                }
            else:
                stats_dict[col] = {
                    'Count': df[col].count(),
                    'Unique': df[col].nunique(),
                    'Mode': df[col].mode()[0] if not df[col].mode().empty else None
                }
        
        stats_df = pd.DataFrame(stats_dict).T
        return stats_df
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None

def create_pivot_table(df, index_col, columns_col, values_col, aggfunc):
    """Create pivot table"""
    try:
        pivot = pd.pivot_table(df, index=index_col, columns=columns_col, 
                              values=values_col, aggfunc=aggfunc, fill_value=0)
        return pivot
    except Exception as e:
        st.error(f"Error creating pivot table: {str(e)}")
        return None

def filter_data(df, column, condition, value):
    """Filter dataframe based on condition"""
    try:
        if condition == "equals":
            return df[df[column] == value]
        elif condition == "contains":
            return df[df[column].astype(str).str.contains(str(value), case=False, na=False)]
        elif condition == "greater than":
            return df[df[column] > float(value)]
        elif condition == "less than":
            return df[df[column] < float(value)]
        elif condition == "not equals":
            return df[df[column] != value]
        else:
            return df
    except Exception as e:
        st.error(f"Error filtering data: {str(e)}")
        return df

def search_in_excel(wb, search_term, case_sensitive=False):
    """Search for term across all sheets"""
    results = []
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        search_value = search_term if case_sensitive else search_term.lower()
                        compare_value = cell_value if case_sensitive else cell_value.lower()
                        
                        if search_value in compare_value:
                            results.append({
                                'Sheet': sheet_name,
                                'Cell': cell.coordinate,
                                'Value': cell_value
                            })
        return pd.DataFrame(results)
    except Exception as e:
        st.error(f"Error searching: {str(e)}")
        return pd.DataFrame()

# ==================== BULK OPERATIONS FUNCTIONS ====================

def batch_modify_cells(wb, modifications_df):
    """Batch modify cells from CSV data"""
    try:
        results = []
        for idx, row in modifications_df.iterrows():
            try:
                cell_address = row['CellAddress']
                new_value = row['NewValue']
                sheet_name = row.get('SheetName', wb.sheetnames[0])
                
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet[cell_address] = new_value
                    results.append({'Row': idx+1, 'Status': 'Success', 'Message': f'Updated {cell_address}'})
                else:
                    results.append({'Row': idx+1, 'Status': 'Error', 'Message': f'Sheet {sheet_name} not found'})
            except Exception as e:
                results.append({'Row': idx+1, 'Status': 'Error', 'Message': str(e)})
        
        return wb, pd.DataFrame(results)
    except Exception as e:
        st.error(f"Error in batch modification: {str(e)}")
        return wb, pd.DataFrame()

def merge_excel_files(file_list, merge_option):
    """Merge multiple Excel files"""
    try:
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        
        sheet_counter = {}
        
        for file_bytes, file_name in file_list:
            file_io = BytesIO(file_bytes)
            wb = load_workbook(file_io)
            
            for sheet_name in wb.sheetnames:
                # Handle duplicate sheet names
                original_name = sheet_name
                counter = sheet_counter.get(original_name, 0)
                
                if counter > 0:
                    new_sheet_name = f"{original_name}_{counter}"
                else:
                    new_sheet_name = original_name
                
                sheet_counter[original_name] = counter + 1
                
                # Copy sheet
                source_sheet = wb[sheet_name]
                target_sheet = new_wb.create_sheet(title=new_sheet_name)
                
                # Copy data
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = copy(cell.number_format)
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)
            
            wb.close()
        
        return new_wb
    except Exception as e:
        st.error(f"Error merging files: {str(e)}")
        return None

def split_excel_by_column(df, split_column, original_filename):
    """Split Excel file by unique values in a column"""
    try:
        unique_values = df[split_column].unique()
        files_dict = {}
        
        for value in unique_values:
            filtered_df = df[df[split_column] == value]
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Write data
            for r in dataframe_to_rows(filtered_df, index=False, header=True):
                ws.append(r)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            safe_value = str(value).replace('/', '_').replace('\\', '_')[:20]
            filename = f"{original_filename}_{safe_value}.xlsx"
            files_dict[filename] = output.getvalue()
        
        return files_dict
    except Exception as e:
        st.error(f"Error splitting file: {str(e)}")
        return {}

def copy_data_between_sheets(wb, source_sheet, source_range, dest_sheet, dest_start):
    """Copy data from one sheet to another"""
    try:
        src_sheet = wb[source_sheet]
        dst_sheet = wb[dest_sheet]
        
        # Parse range (simple implementation for A1:B10 format)
        if ':' in source_range:
            cells = src_sheet[source_range]
            
            # Get starting cell coordinates
            dest_cell = dst_sheet[dest_start]
            start_row = dest_cell.row
            start_col = dest_cell.column
            
            # Copy data
            for i, row in enumerate(cells):
                for j, cell in enumerate(row):
                    target_cell = dst_sheet.cell(row=start_row + i, column=start_col + j)
                    target_cell.value = cell.value
        
        return wb
    except Exception as e:
        st.error(f"Error copying data: {str(e)}")
        return wb

def delete_rows_by_condition(df, column, condition, value):
    """Delete rows based on condition"""
    try:
        if condition == "equals":
            filtered_df = df[df[column] != value]
        elif condition == "contains":
            filtered_df = df[~df[column].astype(str).str.contains(str(value), case=False, na=False)]
        elif condition == "greater than":
            filtered_df = df[df[column] <= float(value)]
        elif condition == "less than":
            filtered_df = df[df[column] >= float(value)]
        elif condition == "empty":
            filtered_df = df[df[column].notna()]
        else:
            filtered_df = df
        
        deleted_count = len(df) - len(filtered_df)
        return filtered_df, deleted_count
    except Exception as e:
        st.error(f"Error deleting rows: {str(e)}")
        return df, 0

def find_and_replace(wb, find_text, replace_text, match_case=False, match_entire=False, sheet_name=None):
    """Find and replace text in workbook"""
    try:
        replacements = []
        sheets_to_search = [sheet_name] if sheet_name else wb.sheetnames
        
        for sname in sheets_to_search:
            sheet = wb[sname]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        search_value = find_text if match_case else find_text.lower()
                        compare_value = cell_value if match_case else cell_value.lower()
                        
                        if match_entire:
                            if compare_value == search_value:
                                cell.value = replace_text
                                replacements.append({
                                    'Sheet': sname,
                                    'Cell': cell.coordinate,
                                    'Old Value': cell_value,
                                    'New Value': replace_text
                                })
                        else:
                            if search_value in compare_value:
                                if match_case:
                                    new_value = cell_value.replace(find_text, replace_text)
                                else:
                                    # Case-insensitive replace
                                    import re
                                    new_value = re.sub(re.escape(find_text), replace_text, cell_value, flags=re.IGNORECASE)
                                
                                cell.value = new_value
                                replacements.append({
                                    'Sheet': sname,
                                    'Cell': cell.coordinate,
                                    'Old Value': cell_value,
                                    'New Value': new_value
                                })
        
        return wb, pd.DataFrame(replacements)
    except Exception as e:
        st.error(f"Error in find and replace: {str(e)}")
        return wb, pd.DataFrame()

# ==================== SHEET MANAGEMENT FUNCTIONS ====================

def add_sheet(wb, sheet_name, position='end'):
    """Add new sheet to workbook"""
    try:
        if position == 'end':
            wb.create_sheet(title=sheet_name)
        elif position == 'beginning':
            wb.create_sheet(title=sheet_name, index=0)
        return wb
    except Exception as e:
        st.error(f"Error adding sheet: {str(e)}")
        return wb

def delete_sheet(wb, sheet_name):
    """Delete sheet from workbook"""
    try:
        if len(wb.sheetnames) > 1:
            del wb[sheet_name]
        else:
            st.error("Cannot delete the last sheet in the workbook")
        return wb
    except Exception as e:
        st.error(f"Error deleting sheet: {str(e)}")
        return wb

def rename_sheet(wb, old_name, new_name):
    """Rename sheet"""
    try:
        sheet = wb[old_name]
        sheet.title = new_name
        return wb
    except Exception as e:
        st.error(f"Error renaming sheet: {str(e)}")
        return wb

def reorder_sheets(wb, new_order):
    """Reorder sheets in workbook"""
    try:
        wb._sheets = [wb[name] for name in new_order]
        return wb
    except Exception as e:
        st.error(f"Error reordering sheets: {str(e)}")
        return wb

def hide_unhide_sheet(wb, sheet_name, hide=True):
    """Hide or unhide a sheet"""
    try:
        sheet = wb[sheet_name]
        if hide:
            # Check if it's the last visible sheet
            visible_count = sum(1 for s in wb.worksheets if s.sheet_state == 'visible')
            if visible_count <= 1:
                st.error("Cannot hide the last visible sheet")
                return wb
            sheet.sheet_state = 'hidden'
        else:
            sheet.sheet_state = 'visible'
        return wb
    except Exception as e:
        st.error(f"Error hiding/unhiding sheet: {str(e)}")
        return wb

def protect_sheet(wb, sheet_name, password=None):
    """Protect a sheet with optional password"""
    try:
        sheet = wb[sheet_name]
        sheet.protection.sheet = True
        if password:
            sheet.protection.password = password
        return wb
    except Exception as e:
        st.error(f"Error protecting sheet: {str(e)}")
        return wb

def unprotect_sheet(wb, sheet_name):
    """Unprotect a sheet"""
    try:
        sheet = wb[sheet_name]
        sheet.protection.sheet = False
        sheet.protection.password = None
        return wb
    except Exception as e:
        st.error(f"Error unprotecting sheet: {str(e)}")
        return wb

# ==================== MAIN APP ====================

st.title("📊 Excel Manipulation Tool - Professional Edition")
st.markdown("---")

# Sidebar
with st.sidebar:
    st.header("🎯 Navigation")
    st.info("Upload an Excel file to unlock all features")
    
    if st.session_state.uploaded_file:
        st.success("✅ File loaded successfully")
        if st.session_state.df_dict:
            st.metric("Sheets", len(st.session_state.df_dict))

# Main tabs
tab1, tab2, tab3, tab4 = st.tabs([
    "📁 Basic Operations", 
    "📊 Data Analysis & Visualization", 
    "⚡ Bulk Operations", 
    "📑 Sheet Management"
])

# ==================== TAB 1: BASIC OPERATIONS ====================
with tab1:
    st.header("📁 Basic Operations")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Create New Excel File")
        file_name = st.text_input("Enter file name:", key="new_file_name")
        if st.button("Create File", key="create_btn"):
            if file_name:
                create_new_excel(file_name)
            else:
                st.warning("Please enter a file name")
    
    with col2:
        st.subheader("Upload Excel File")
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"], key="file_uploader")
        password = st.text_input("Password (if protected):", type="password", key="file_password")
    
    if uploaded_file is not None:
        # Save to session state
        st.session_state.uploaded_file = uploaded_file
        
        # Load file
        file_bytes = uploaded_file.getvalue()
        file_io = load_excel_with_password(file_bytes, password if password else None)
        
        if file_io:
            # Save to temp file for operations
            temp_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
            with open(temp_path, 'wb') as f:
                f.write(file_bytes)
            st.session_state.file_path = temp_path
            
            # Load workbook
            try:
                file_io.seek(0)
                wb = load_workbook(file_io)
                st.session_state.workbook = wb
                
                # Load all sheets
                file_io.seek(0)
                sheets = get_all_sheets(file_io)
                st.session_state.df_dict = {}
                
                for sheet in sheets:
                    file_io.seek(0)
                    df = load_sheet_data(file_io, sheet)
                    if df is not None:
                        st.session_state.df_dict[sheet] = df
                
                st.success(f"✅ Loaded {len(sheets)} sheet(s)")
                
                # Display data
                st.subheader("📋 Preview Data")
                selected_sheet = st.selectbox("Select sheet to view:", sheets, key="preview_sheet")
                
                if selected_sheet in st.session_state.df_dict:
                    df = st.session_state.df_dict[selected_sheet]
                    st.dataframe(df.head(100), use_container_width=True)
                    st.info(f"Showing first 100 rows of {len(df)} total rows")
                
                # Modify cell
                st.subheader("✏️ Modify Cell")
                mod_col1, mod_col2, mod_col3 = st.columns(3)
                
                with mod_col1:
                    mod_sheet = st.selectbox("Sheet:", sheets, key="mod_sheet")
                with mod_col2:
                    cell_address = st.text_input("Cell address (e.g., A1):", key="cell_addr")
                with mod_col3:
                    new_value = st.text_input("New value:", key="new_val")
                
                if st.button("Modify Cell", key="modify_cell_btn"):
                    if cell_address and new_value:
                        wb = modify_excel_cell(st.session_state.workbook, mod_sheet, cell_address.upper(), new_value)
                        st.session_state.workbook = wb
                        
                        # Provide download
                        st.download_button(
                            label="📥 Download Modified File",
                            data=create_download_link(wb, uploaded_file.name),
                            file_name=f"modified_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # Password operations
                st.subheader("🔒 Password Operations")
                pw_col1, pw_col2 = st.columns(2)
                
                with pw_col1:
                    st.write("**Set Password**")
                    new_password = st.text_input("New password:", type="password", key="set_pw")
                    if st.button("Set Password", key="set_pw_btn"):
                        if new_password and st.session_state.file_path:
                            set_password_excel(st.session_state.file_path, new_password)
                
                with pw_col2:
                    st.write("**Remove Password**")
                    remove_pw = st.text_input("Current password:", type="password", key="remove_pw")
                    if st.button("Remove Password", key="remove_pw_btn"):
                        if remove_pw and st.session_state.file_path:
                            remove_password_excel(st.session_state.file_path, remove_pw)
                
            except Exception as e:
                st.error(f"Error loading workbook: {str(e)}")

# ==================== TAB 2: DATA ANALYSIS ====================
with tab2:
    st.header("📊 Data Analysis & Visualization")
    
    if not st.session_state.df_dict:
        st.warning("⚠️ Please upload an Excel file in the Basic Operations tab first")
    else:
        # Chart Generation
        with st.expander("📈 Chart Generation", expanded=True):
            chart_sheet = st.selectbox("Select sheet:", list(st.session_state.df_dict.keys()), key="chart_sheet")
            df = st.session_state.df_dict[chart_sheet]
            
            chart_col1, chart_col2, chart_col3 = st.columns(3)
            
            with chart_col1:
                chart_type = st.selectbox("Chart type:", ["Bar Chart", "Line Chart", "Pie Chart", "Scatter Plot"], key="chart_type")
            with chart_col2:
                x_column = st.selectbox("X-axis / Names:", df.columns.tolist(), key="x_col")
            with chart_col3:
                y_column = st.selectbox("Y-axis / Values:", df.columns.tolist(), key="y_col")
            
            chart_title = st.text_input("Chart title:", value=f"{chart_type} - {y_column} by {x_column}", key="chart_title")
            
            if st.button("Generate Chart", key="gen_chart"):
                fig = create_chart(df, chart_type, x_column, y_column, chart_title)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
        
        # Statistics
        with st.expander("📊 Statistical Calculations"):
            stats_sheet = st.selectbox("Select sheet:", list(st.session_state.df_dict.keys()), key="stats_sheet")
            df = st.session_state.df_dict[stats_sheet]
            
            selected_columns = st.multiselect("Select columns:", df.columns.tolist(), key="stats_cols")
            
            if st.button("Calculate Statistics", key="calc_stats"):
                if selected_columns:
                    stats_df = calculate_statistics(df, selected_columns)
                    if stats_df is not None:
                        st.dataframe(stats_df, use_container_width=True)
                        
                        # Option to save to new sheet
                        if st.button("Save Statistics to New Sheet", key="save_stats"):
                            wb = st.session_state.workbook
                            new_sheet = wb.create_sheet(title="Statistics")
                            
                            for r in dataframe_to_rows(stats_df, index=True, header=True):
                                new_sheet.append(r)
                            
                            st.download_button(
                                label="📥 Download with Statistics",
                                data=create_download_link(wb, "statistics.xlsx"),
                                file_name="excel_with_statistics.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                else:
                    st.warning("Please select at least one column")
        
        # Pivot Table
        with st.expander("🔄 Pivot Table Creator"):
            pivot_sheet = st.selectbox("Select sheet:", list(st.session_state.df_dict.keys()), key="pivot_sheet")
            df = st.session_state.df_dict[pivot_sheet]
            
            piv_col1, piv_col2, piv_col3, piv_col4 = st.columns(4)
            
            with piv_col1:
                index_col = st.selectbox("Rows (Index):", df.columns.tolist(), key="pivot_index")
            with piv_col2:
                columns_col = st.selectbox("Columns:", df.columns.tolist(), key="pivot_cols")
            with piv_col3:
                values_col = st.selectbox("Values:", df.columns.tolist(), key="pivot_vals")
            with piv_col4:
                aggfunc = st.selectbox("Aggregation:", ["sum", "mean", "count", "min", "max"], key="pivot_agg")
            
            if st.button("Create Pivot Table", key="create_pivot"):
                pivot_df = create_pivot_table(df, index_col, columns_col, values_col, aggfunc)
                if pivot_df is not None:
                    st.dataframe(pivot_df, use_container_width=True)
                    
                    # Save option
                    if st.button("Save Pivot to New Sheet", key="save_pivot"):
                        wb = st.session_state.workbook
                        new_sheet = wb.create_sheet(title="Pivot_Table")
                        
                        for r in dataframe_to_rows(pivot_df, index=True, header=True):
                            new_sheet.append(r)
                        
                        st.download_button(
                            label="📥 Download with Pivot Table",
                            data=create_download_link(wb, "pivot.xlsx"),
                            file_name="excel_with_pivot.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        
        # Filter & Sort
        with st.expander("🔍 Filter & Sort Data"):
            filter_sheet = st.selectbox("Select sheet:", list(st.session_state.df_dict.keys()), key="filter_sheet")
            df = st.session_state.df_dict[filter_sheet]
            
            st.write("**Filter Options**")
            filt_col1, filt_col2, filt_col3 = st.columns(3)
            
            with filt_col1:
                filter_column = st.selectbox("Column:", df.columns.tolist(), key="filter_col")
            with filt_col2:
                filter_condition = st.selectbox("Condition:", ["equals", "contains", "greater than", "less than", "not equals"], key="filter_cond")
            with filt_col3:
                filter_value = st.text_input("Value:", key="filter_val")
            
            st.write("**Sort Options**")
            sort_col1, sort_col2 = st.columns(2)
            
            with sort_col1:
                sort_column = st.selectbox("Sort by:", df.columns.tolist(), key="sort_col")
            with sort_col2:
                sort_order = st.radio("Order:", ["Ascending", "Descending"], key="sort_order", horizontal=True)
            
            if st.button("Apply Filter & Sort", key="apply_filter"):
                result_df = df.copy()
                
                # Apply filter
                if filter_value:
                    result_df = filter_data(result_df, filter_column, filter_condition, filter_value)
                
                # Apply sort
                result_df = result_df.sort_values(by=sort_column, ascending=(sort_order == "Ascending"))
                
                st.success(f"Filtered to {len(result_df)} rows from {len(df)} total rows")
                st.dataframe(result_df.head(100), use_container_width=True)
                
                # Save option
                if st.button("Save Filtered Data", key="save_filtered"):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Filtered_Data"
                    
                    for r in dataframe_to_rows(result_df, index=False, header=True):
                        ws.append(r)
                    
                    st.download_button(
                        label="📥 Download Filtered Data",
                        data=create_download_link(wb, "filtered.xlsx"),
                        file_name="filtered_data.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # Search
        with st.expander("🔎 Search Functionality"):
            search_term = st.text_input("Search for:", key="search_term")
            case_sensitive = st.checkbox("Case sensitive", key="case_sens")
            
            if st.button("Search", key="search_btn"):
                if search_term and st.session_state.workbook:
                    with st.spinner("Searching..."):
                        results_df = search_in_excel(st.session_state.workbook, search_term, case_sensitive)
                    
                    if not results_df.empty:
                        st.success(f"Found {len(results_df)} matches")
                        st.dataframe(results_df, use_container_width=True)
                    else:
                        st.info("No matches found")
                else:
                    st.warning("Please enter a search term")

# ==================== TAB 3: BULK OPERATIONS ====================
with tab3:
    st.header("⚡ Bulk Operations")
    
    if not st.session_state.workbook:
        st.warning("⚠️ Please upload an Excel file in the Basic Operations tab first")
    else:
        # Batch Modify
        with st.expander("📝 Batch Modify Multiple Cells"):
            st.info("Upload a CSV file with columns: CellAddress, NewValue, SheetName (optional)")
            
            batch_csv = st.file_uploader("Upload CSV file:", type=["csv"], key="batch_csv")
            
            if batch_csv:
                try:
                    modifications_df = pd.read_csv(batch_csv)
                    st.write("**Preview of modifications:**")
                    st.dataframe(modifications_df.head(10))
                    
                    if st.button("Apply Batch Modifications", key="apply_batch"):
                        with st.spinner("Applying modifications..."):
                            wb, results_df = batch_modify_cells(st.session_state.workbook, modifications_df)
                            st.session_state.workbook = wb
                        
                        st.write("**Results:**")
                        st.dataframe(results_df)
                        
                        st.download_button(
                            label="📥 Download Modified File",
                            data=create_download_link(wb, "batch_modified.xlsx"),
                            file_name="batch_modified.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Error reading CSV: {str(e)}")
        
        # Merge Files
        with st.expander("🔗 Merge Multiple Excel Files"):
            merge_files = st.file_uploader("Upload Excel files to merge:", type=["xlsx"], accept_multiple_files=True, key="merge_files")
            
            if merge_files and len(merge_files) > 1:
                st.info(f"Selected {len(merge_files)} files to merge")
                
                if st.button("Merge Files", key="merge_btn"):
                    with st.spinner("Merging files..."):
                        file_list = [(f.getvalue(), f.name) for f in merge_files]
                        merged_wb = merge_excel_files(file_list, "all_sheets")
                    
                    if merged_wb:
                        st.success(f"✅ Merged into {len(merged_wb.sheetnames)} sheets")
                        
                        st.download_button(
                            label="📥 Download Merged File",
                            data=create_download_link(merged_wb, "merged.xlsx"),
                            file_name="merged_excel.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            elif merge_files:
                st.warning("Please upload at least 2 files to merge")
        
        # Split File
        with st.expander("✂️ Split Excel File by Criteria"):
            if st.session_state.df_dict:
                split_sheet = st.selectbox("Select sheet to split:", list(st.session_state.df_dict.keys()), key="split_sheet")
                df = st.session_state.df_dict[split_sheet]
                
                split_column = st.selectbox("Split by column:", df.columns.tolist(), key="split_col")
                
                if split_column:
                    unique_values = df[split_column].unique()
                    st.info(f"This will create {len(unique_values)} separate files")
                
                if st.button("Split File", key="split_btn"):
                    with st.spinner("Splitting file..."):
                        files_dict = split_excel_by_column(df, split_column, "split")
                    
                    if files_dict:
                        # Create zip file
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for filename, file_data in files_dict.items():
                                zip_file.writestr(filename, file_data)
                        
                        zip_buffer.seek(0)
                        
                        st.success(f"✅ Created {len(files_dict)} files")
                        st.download_button(
                            label="📥 Download All Files (ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name="split_files.zip",
                            mime="application/zip"
                        )
        
        # Copy Data Between Sheets
        with st.expander("📋 Copy Data Between Sheets"):
            wb = st.session_state.workbook
            sheets = wb.sheetnames
            
            copy_col1, copy_col2 = st.columns(2)
            
            with copy_col1:
                st.write("**Source**")
                source_sheet = st.selectbox("Source sheet:", sheets, key="copy_src_sheet")
                source_range = st.text_input("Source range (e.g., A1:C10):", key="copy_src_range")
            
            with copy_col2:
                st.write("**Destination**")
                dest_sheet = st.selectbox("Destination sheet:", sheets, key="copy_dest_sheet")
                dest_start = st.text_input("Destination start cell (e.g., A1):", key="copy_dest_start")
            
            if st.button("Copy Data", key="copy_data_btn"):
                if source_range and dest_start:
                    wb = copy_data_between_sheets(wb, source_sheet, source_range, dest_sheet, dest_start)
                    st.session_state.workbook = wb
                    
                    st.success("✅ Data copied successfully")
                    st.download_button(
                        label="📥 Download Updated File",
                        data=create_download_link(wb, "copied.xlsx"),
                        file_name="data_copied.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Please fill in all fields")
        
        # Delete Rows by Condition
        with st.expander("🗑️ Delete Rows by Condition"):
            if st.session_state.df_dict:
                del_sheet = st.selectbox("Select sheet:", list(st.session_state.df_dict.keys()), key="del_sheet")
                df = st.session_state.df_dict[del_sheet]
                
                del_col1, del_col2, del_col3 = st.columns(3)
                
                with del_col1:
                    del_column = st.selectbox("Column:", df.columns.tolist(), key="del_col")
                with del_col2:
                    del_condition = st.selectbox("Condition:", ["equals", "contains", "greater than", "less than", "empty"], key="del_cond")
                with del_col3:
                    del_value = st.text_input("Value:", key="del_val")
                
                if st.button("Preview Deletion", key="preview_del"):
                    filtered_df, deleted_count = delete_rows_by_condition(df, del_column, del_condition, del_value)
                    
                    st.warning(f"⚠️ This will delete {deleted_count} rows")
                    st.write("**Remaining data preview:**")
                    st.dataframe(filtered_df.head(20))
                    
                    if st.button("Confirm Deletion", key="confirm_del"):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = del_sheet
                        
                        for r in dataframe_to_rows(filtered_df, index=False, header=True):
                            ws.append(r)
                        
                        st.success(f"✅ Deleted {deleted_count} rows")
                        st.download_button(
                            label="📥 Download Updated File",
                            data=create_download_link(wb, "deleted.xlsx"),
                            file_name="rows_deleted.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        
        # Find and Replace
        with st.expander("🔄 Find and Replace"):
            wb = st.session_state.workbook
            
            find_col1, find_col2 = st.columns(2)
            
            with find_col1:
                find_text = st.text_input("Find:", key="find_text")
            with find_col2:
                replace_text = st.text_input("Replace with:", key="replace_text")
            
            opt_col1, opt_col2, opt_col3 = st.columns(3)
            
            with opt_col1:
                match_case = st.checkbox("Match case", key="match_case")
            with opt_col2:
                match_entire = st.checkbox("Match entire cell", key="match_entire")
            with opt_col3:
                search_sheet = st.selectbox("Search in:", ["All sheets"] + wb.sheetnames, key="search_sheet")
            
            if st.button("Preview Replacements", key="preview_replace"):
                if find_text:
                    sheet_name = None if search_sheet == "All sheets" else search_sheet
                    
                    # Create a copy for preview
                    wb_copy = load_workbook(BytesIO(create_download_link(wb, "temp.xlsx")))
                    _, replacements_df = find_and_replace(wb_copy, find_text, replace_text, match_case, match_entire, sheet_name)
                    
                    if not replacements_df.empty:
                        st.info(f"Found {len(replacements_df)} matches")
                        st.dataframe(replacements_df.head(50))
                        
                        if st.button("Confirm Replace", key="confirm_replace"):
                            wb, _ = find_and_replace(wb, find_text, replace_text, match_case, match_entire, sheet_name)
                            st.session_state.workbook = wb
                            
                            st.success(f"✅ Replaced {len(replacements_df)} occurrences")
                            st.download_button(
                                label="📥 Download Updated File",
                                data=create_download_link(wb, "replaced.xlsx"),
                                file_name="find_replace.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.info("No matches found")
                else:
                    st.warning("Please enter text to find")

# ==================== TAB 4: SHEET MANAGEMENT ====================
with tab4:
    st.header("📑 Sheet Management")
    
    if not st.session_state.workbook:
        st.warning("⚠️ Please upload an Excel file in the Basic Operations tab first")
    else:
        wb = st.session_state.workbook
        
        # Add/Delete/Rename Sheets
        with st.expander("➕ Add / ✏️ Rename / 🗑️ Delete Sheets", expanded=True):
            sheet_col1, sheet_col2, sheet_col3 = st.columns(3)
            
            with sheet_col1:
                st.write("**Add Sheet**")
                new_sheet_name = st.text_input("New sheet name:", key="new_sheet_name")
                position = st.radio("Position:", ["End", "Beginning"], key="sheet_position", horizontal=True)
                
                if st.button("Add Sheet", key="add_sheet_btn"):
                    if new_sheet_name:
                        valid, msg = validate_sheet_name(new_sheet_name, wb.sheetnames)
                        if valid:
                            wb = add_sheet(wb, new_sheet_name, position.lower())
                            st.session_state.workbook = wb
                            st.success(f"✅ Added sheet '{new_sheet_name}'")
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Please enter a sheet name")
            
            with sheet_col2:
                st.write("**Rename Sheet**")
                old_sheet_name = st.selectbox("Select sheet:", wb.sheetnames, key="rename_old")
                rename_new_name = st.text_input("New name:", key="rename_new")
                
                if st.button("Rename Sheet", key="rename_sheet_btn"):
                    if rename_new_name:
                        valid, msg = validate_sheet_name(rename_new_name, [s for s in wb.sheetnames if s != old_sheet_name])
                        if valid:
                            wb = rename_sheet(wb, old_sheet_name, rename_new_name)
                            st.session_state.workbook = wb
                            st.success(f"✅ Renamed to '{rename_new_name}'")
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Please enter a new name")
            
            with sheet_col3:
                st.write("**Delete Sheet**")
                delete_sheet_name = st.selectbox("Select sheet:", wb.sheetnames, key="delete_sheet")
                
                if st.button("Delete Sheet", key="delete_sheet_btn"):
                    if len(wb.sheetnames) > 1:
                        wb = delete_sheet(wb, delete_sheet_name)
                        st.session_state.workbook = wb
                        st.success(f"✅ Deleted sheet '{delete_sheet_name}'")
                        st.rerun()
                    else:
                        st.error("Cannot delete the last sheet")
            
            # Download button for changes
            if st.button("💾 Save All Sheet Changes", key="save_sheet_changes"):
                st.download_button(
                    label="📥 Download Updated File",
                    data=create_download_link(wb, "updated.xlsx"),
                    file_name="sheets_updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # Reorder Sheets
        with st.expander("🔀 Reorder Sheets"):
            st.write("**Current order:**")
            current_order = wb.sheetnames
            
            for i, sheet in enumerate(current_order):
                st.write(f"{i+1}. {sheet}")
            
            st.write("**New order (comma-separated):**")
            st.info("Enter sheet names in desired order, separated by commas")
            new_order_input = st.text_input("New order:", value=", ".join(current_order), key="new_order")
            
            if st.button("Apply New Order", key="reorder_btn"):
                new_order = [s.strip() for s in new_order_input.split(",")]
                
                if set(new_order) == set(current_order):
                    wb = reorder_sheets(wb, new_order)
                    st.session_state.workbook = wb
                    st.success("✅ Sheets reordered successfully")
                    
                    st.download_button(
                        label="📥 Download Reordered File",
                        data=create_download_link(wb, "reordered.xlsx"),
                        file_name="sheets_reordered.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Invalid order. Please include all sheets exactly once.")
        
        # Hide/Unhide Sheets
        with st.expander("👁️ Hide / Unhide Sheets"):
            st.write("**Sheet Visibility Status:**")
            
            for sheet in wb.worksheets:
                col1, col2, col3 = st.columns([3, 1, 1])
                
                with col1:
                    st.write(f"**{sheet.title}**")
                with col2:
                    status = "Visible" if sheet.sheet_state == 'visible' else "Hidden"
                    st.write(status)
                with col3:
                    if sheet.sheet_state == 'visible':
                        if st.button("Hide", key=f"hide_{sheet.title}"):
                            wb = hide_unhide_sheet(wb, sheet.title, hide=True)
                            st.session_state.workbook = wb
                            st.rerun()
                    else:
                        if st.button("Unhide", key=f"unhide_{sheet.title}"):
                            wb = hide_unhide_sheet(wb, sheet.title, hide=False)
                            st.session_state.workbook = wb
                            st.rerun()
            
            if st.button("💾 Save Visibility Changes", key="save_visibility"):
                st.download_button(
                    label="📥 Download Updated File",
                    data=create_download_link(wb, "visibility.xlsx"),
                    file_name="visibility_updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # Protect/Unprotect Sheets
        with st.expander("🔒 Protect / Unprotect Sheets"):
            st.write("**Sheet Protection Status:**")
            
            for sheet in wb.worksheets:
                st.write(f"**{sheet.title}**")
                
                prot_col1, prot_col2 = st.columns(2)
                
                with prot_col1:
                    is_protected = sheet.protection.sheet
                    st.write(f"Status: {'🔒 Protected' if is_protected else '🔓 Unprotected'}")
                
                with prot_col2:
                    if not is_protected:
                        protect_pw = st.text_input(f"Password for {sheet.title}:", type="password", key=f"protect_pw_{sheet.title}")
                        if st.button(f"Protect", key=f"protect_{sheet.title}"):
                            wb = protect_sheet(wb, sheet.title, protect_pw if protect_pw else None)
                            st.session_state.workbook = wb
                            st.success(f"✅ Protected '{sheet.title}'")
                            st.rerun()
                    else:
                        if st.button(f"Unprotect", key=f"unprotect_{sheet.title}"):
                            wb = unprotect_sheet(wb, sheet.title)
                            st.session_state.workbook = wb
                            st.success(f"✅ Unprotected '{sheet.title}'")
                            st.rerun()
                
                st.markdown("---")
            
            if st.button("💾 Save Protection Changes", key="save_protection"):
                st.download_button(
                    label="📥 Download Updated File",
                    data=create_download_link(wb, "protection.xlsx"),
                    file_name="protection_updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <p>Excel Manipulation Tool - Professional Edition | Built with Streamlit & Python</p>
</div>
""", unsafe_allow_html=True)
