"""
Bulk Operations and Automation
Functions for batch modifications, merging, splitting, copying, deleting, and find/replace
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from copy import copy
import re


def batch_modify_cells(wb, modifications_df):
    """
    Batch modify cells from CSV data
    
    Args:
        wb: openpyxl Workbook object
        modifications_df: DataFrame with columns: CellAddress, NewValue, SheetName (optional)
        
    Returns:
        Tuple of (modified workbook, results DataFrame)
    """
    try:
        results = []
        for idx, row in modifications_df.iterrows():
            try:
                cell_address = row['CellAddress']
                new_value = row['NewValue']
                sheet_name = row.get('SheetName', wb.sheetnames[0])
                
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet[cell_address] = new_value
                    results.append({'Row': idx+1, 'Status': 'Success', 'Message': f'Updated {cell_address}'})
                else:
                    results.append({'Row': idx+1, 'Status': 'Error', 'Message': f'Sheet {sheet_name} not found'})
            except Exception as e:
                results.append({'Row': idx+1, 'Status': 'Error', 'Message': str(e)})
        
        return wb, pd.DataFrame(results)
    except Exception as e:
        st.error(f"Error in batch modification: {str(e)}")
        return wb, pd.DataFrame()


def merge_excel_files(file_list, merge_option):
    """
    Merge multiple Excel files into one workbook
    
    Args:
        file_list: List of tuples (file_bytes, file_name)
        merge_option: Merge strategy (currently 'all_sheets')
        
    Returns:
        Merged Workbook object or None on error
    """
    try:
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        
        sheet_counter = {}
        
        for file_bytes, file_name in file_list:
            file_io = BytesIO(file_bytes)
            wb = load_workbook(file_io)
            
            for sheet_name in wb.sheetnames:
                # Handle duplicate sheet names
                original_name = sheet_name
                counter = sheet_counter.get(original_name, 0)
                
                if counter > 0:
                    new_sheet_name = f"{original_name}_{counter}"
                else:
                    new_sheet_name = original_name
                
                sheet_counter[original_name] = counter + 1
                
                # Copy sheet
                source_sheet = wb[sheet_name]
                target_sheet = new_wb.create_sheet(title=new_sheet_name)
                
                # Copy data and formatting
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = copy(cell.number_format)
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)
            
            wb.close()
        
        return new_wb
    except Exception as e:
        st.error(f"Error merging files: {str(e)}")
        return None


def split_excel_by_column(df, split_column, original_filename):
    """
    Split Excel file by unique values in a column
    
    Args:
        df: pandas DataFrame
        split_column: Column name to split by
        original_filename: Base filename for output files
        
    Returns:
        Dictionary of {filename: file_bytes} or empty dict on error
    """
    try:
        unique_values = df[split_column].unique()
        files_dict = {}
        
        for value in unique_values:
            filtered_df = df[df[split_column] == value]
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Write data
            for r in dataframe_to_rows(filtered_df, index=False, header=True):
                ws.append(r)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            safe_value = str(value).replace('/', '_').replace('\\', '_')[:20]
            filename = f"{original_filename}_{safe_value}.xlsx"
            files_dict[filename] = output.getvalue()
        
        return files_dict
    except Exception as e:
        st.error(f"Error splitting file: {str(e)}")
        return {}


def copy_data_between_sheets(wb, source_sheet, source_range, dest_sheet, dest_start):
    """
    Copy data from one sheet to another
    
    Args:
        wb: openpyxl Workbook object
        source_sheet: Source sheet name
        source_range: Cell range (e.g., 'A1:C10')
        dest_sheet: Destination sheet name
        dest_start: Starting cell in destination (e.g., 'A1')
        
    Returns:
        Modified workbook
    """
    try:
        src_sheet = wb[source_sheet]
        dst_sheet = wb[dest_sheet]
        
        # Parse range (simple implementation for A1:B10 format)
        if ':' in source_range:
            cells = src_sheet[source_range]
            
            # Get starting cell coordinates
            dest_cell = dst_sheet[dest_start]
            start_row = dest_cell.row
            start_col = dest_cell.column
            
            # Copy data
            for i, row in enumerate(cells):
                for j, cell in enumerate(row):
                    target_cell = dst_sheet.cell(row=start_row + i, column=start_col + j)
                    target_cell.value = cell.value
        
        return wb
    except Exception as e:
        st.error(f"Error copying data: {str(e)}")
        return wb


def delete_rows_by_condition(df, column, condition, value):
    """
    Delete rows based on condition
    
    Args:
        df: pandas DataFrame
        column: Column name to check
        condition: Condition type (equals, contains, greater than, less than, empty)
        value: Value to compare against
        
    Returns:
        Tuple of (filtered DataFrame, deleted count)
    """
    try:
        if condition == "equals":
            filtered_df = df[df[column] != value]
        elif condition == "contains":
            filtered_df = df[~df[column].astype(str).str.contains(str(value), case=False, na=False)]
        elif condition == "greater than":
            filtered_df = df[df[column] <= float(value)]
        elif condition == "less than":
            filtered_df = df[df[column] >= float(value)]
        elif condition == "empty":
            filtered_df = df[df[column].notna()]
        else:
            filtered_df = df
        
        deleted_count = len(df) - len(filtered_df)
        return filtered_df, deleted_count
    except Exception as e:
        st.error(f"Error deleting rows: {str(e)}")
        return df, 0


def find_and_replace(wb, find_text, replace_text, match_case=False, match_entire=False, sheet_name=None):
    """
    Find and replace text in workbook
    
    Args:
        wb: openpyxl Workbook object
        find_text: Text to find
        replace_text: Replacement text
        match_case: Whether to match case
        match_entire: Whether to match entire cell
        sheet_name: Specific sheet name or None for all sheets
        
    Returns:
        Tuple of (modified workbook, replacements DataFrame)
    """
    try:
        replacements = []
        sheets_to_search = [sheet_name] if sheet_name else wb.sheetnames
        
        for sname in sheets_to_search:
            sheet = wb[sname]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        search_value = find_text if match_case else find_text.lower()
                        compare_value = cell_value if match_case else cell_value.lower()
                        
                        if match_entire:
                            if compare_value == search_value:
                                cell.value = replace_text
                                replacements.append({
                                    'Sheet': sname,
                                    'Cell': cell.coordinate,
                                    'Old Value': cell_value,
                                    'New Value': replace_text
                                })
                        else:
                            if search_value in compare_value:
                                if match_case:
                                    new_value = cell_value.replace(find_text, replace_text)
                                else:
                                    # Case-insensitive replace
                                    new_value = re.sub(re.escape(find_text), replace_text, cell_value, flags=re.IGNORECASE)
                                
                                cell.value = new_value
                                replacements.append({
                                    'Sheet': sname,
                                    'Cell': cell.coordinate,
                                    'Old Value': cell_value,
                                    'New Value': new_value
                                })
        
        return wb, pd.DataFrame(replacements)
    except Exception as e:
        st.error(f"Error in find and replace: {str(e)}")
        return wb, pd.DataFrame()
