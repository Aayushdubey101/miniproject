"""
File Handling Utilities
Functions for loading, saving, and managing Excel files
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import msoffcrypto
from io import BytesIO


@st.cache_data
def load_excel_with_password(file_bytes, password=None):
    """
    Load Excel file with optional password protection
    
    Args:
        file_bytes: File content as bytes
        password: Optional password string
        
    Returns:
        BytesIO object containing decrypted file or None on error
    """
    try:
        if password:
            decrypted = BytesIO()
            file_io = BytesIO(file_bytes)
            office_file = msoffcrypto.OfficeFile(file_io)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return decrypted
        else:
            return BytesIO(file_bytes)
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None


def get_all_sheets(file_io):
    """
    Get all sheet names from Excel file
    
    Args:
        file_io: BytesIO object containing Excel file
        
    Returns:
        List of sheet names or empty list on error
    """
    try:
        wb = load_workbook(file_io, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        st.error(f"Error reading sheets: {str(e)}")
        return []


def load_sheet_data(file_io, sheet_name=None):
    """
    Load data from specific sheet into DataFrame
    
    Args:
        file_io: BytesIO object containing Excel file
        sheet_name: Name of sheet to load (None for first sheet)
        
    Returns:
        pandas DataFrame or None on error
    """
    try:
        df = pd.read_excel(file_io, sheet_name=sheet_name, engine='openpyxl')
        return df
    except Exception as e:
        st.error(f"Error loading sheet data: {str(e)}")
        return None


def create_download_link(wb, filename):
    """
    Create downloadable bytes from workbook
    
    Args:
        wb: openpyxl Workbook object
        filename: Suggested filename (not used, kept for compatibility)
        
    Returns:
        Bytes content of the workbook
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
