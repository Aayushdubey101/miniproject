"""
Tab 3: Bulk Operations UI
Batch modifications, merge, split, copy, delete, and find/replace
"""

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import zipfile
from src.features.bulk_operations import (
    batch_modify_cells, merge_excel_files, split_excel_by_column,
    copy_data_between_sheets, delete_rows_by_condition, find_and_replace
)
from src.utils.file_handlers import create_download_link, load_excel_with_password
from src.ui.components import show_dataframe_preview
from src.config.settings import SESSION_WORKBOOK, SESSION_DF_DICT


def render_bulk_operations_tab():
    """Render the Bulk Operations tab"""
    st.header("⚡ Bulk Operations")
    
    if not st.session_state.get(SESSION_WORKBOOK):
        st.warning("⚠️ Please upload an Excel file in the Basic Operations tab first")
        return
    
    # Batch Modify
    with st.expander("📝 Batch Modify Multiple Cells"):
        st.info("Upload a CSV file with columns: CellAddress, NewValue, SheetName (optional)")
        batch_csv = st.file_uploader("Upload CSV file:", type=["csv"], key="batch_csv")
        
        if batch_csv:
            try:
                modifications_df = pd.read_csv(batch_csv)
                st.write("**Preview of modifications:**")
                st.dataframe(modifications_df.head(10))
                
                if st.button("Apply Batch Modifications", key="apply_batch"):
                    with st.spinner("Applying modifications..."):
                        wb, results_df = batch_modify_cells(st.session_state[SESSION_WORKBOOK], modifications_df)
                        st.session_state[SESSION_WORKBOOK] = wb
                    
                    st.write("**Results:**")
                    st.dataframe(results_df)
                    st.download_button(
                        label="📥 Download Modified File",
                        data=create_download_link(wb, "batch_modified.xlsx"),
                        file_name="batch_modified.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Error reading CSV: {str(e)}")
    
    # Merge Files
    with st.expander("🔗 Merge Multiple Excel Files"):
        merge_files = st.file_uploader("Upload Excel files to merge:", type=["xlsx"], accept_multiple_files=True, key="merge_files")
        
        if merge_files and len(merge_files) > 1:
            st.info(f"Selected {len(merge_files)} files to merge")
            
            if st.button("Merge Files", key="merge_btn"):
                with st.spinner("Merging files..."):
                    file_list = [(f.getvalue(), f.name) for f in merge_files]
                    merged_wb = merge_excel_files(file_list, "all_sheets")
                
                if merged_wb:
                    st.success(f"✅ Merged into {len(merged_wb.sheetnames)} sheets")
                    st.download_button(
                        label="📥 Download Merged File",
                        data=create_download_link(merged_wb, "merged.xlsx"),
                        file_name="merged_excel.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        elif merge_files:
            st.warning("Please upload at least 2 files to merge")
    
    # Split File
    with st.expander("✂️ Split Excel File by Criteria"):
        if st.session_state.get(SESSION_DF_DICT):
            split_sheet = st.selectbox("Select sheet to split:", list(st.session_state[SESSION_DF_DICT].keys()), key="split_sheet")
            df = st.session_state[SESSION_DF_DICT][split_sheet]
            
            split_column = st.selectbox("Split by column:", df.columns.tolist(), key="split_col")
            
            if split_column:
                unique_values = df[split_column].unique()
                st.info(f"This will create {len(unique_values)} separate files")
            
            if st.button("Split File", key="split_btn"):
                with st.spinner("Splitting file..."):
                    files_dict = split_excel_by_column(df, split_column, "split")
                
                if files_dict:
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for filename, file_data in files_dict.items():
                            zip_file.writestr(filename, file_data)
                    zip_buffer.seek(0)
                    
                    st.success(f"✅ Created {len(files_dict)} files")
                    st.download_button(
                        label="📥 Download All Files (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="split_files.zip",
                        mime="application/zip"
                    )
    
    # Copy Data Between Sheets
    with st.expander("📋 Copy Data Between Sheets"):
        wb = st.session_state[SESSION_WORKBOOK]
        sheets = wb.sheetnames
        
        copy_col1, copy_col2 = st.columns(2)
        with copy_col1:
            st.write("**Source**")
            source_sheet = st.selectbox("Source sheet:", sheets, key="copy_src_sheet")
            source_range = st.text_input("Source range (e.g., A1:C10):", key="copy_src_range")
        with copy_col2:
            st.write("**Destination**")
            dest_sheet = st.selectbox("Destination sheet:", sheets, key="copy_dest_sheet")
            dest_start = st.text_input("Destination start cell (e.g., A1):", key="copy_dest_start")
        
        if st.button("Copy Data", key="copy_data_btn"):
            if source_range and dest_start:
                wb = copy_data_between_sheets(wb, source_sheet, source_range, dest_sheet, dest_start)
                st.session_state[SESSION_WORKBOOK] = wb
                st.success("✅ Data copied successfully")
                st.download_button(
                    label="📥 Download Updated File",
                    data=create_download_link(wb, "copied.xlsx"),
                    file_name="data_copied.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Please fill in all fields")
    
    # Delete Rows by Condition
    with st.expander("🗑️ Delete Rows by Condition"):
        if st.session_state.get(SESSION_DF_DICT):
            del_sheet = st.selectbox("Select sheet:", list(st.session_state[SESSION_DF_DICT].keys()), key="del_sheet")
            df = st.session_state[SESSION_DF_DICT][del_sheet]
            
            del_col1, del_col2, del_col3 = st.columns(3)
            with del_col1:
                del_column = st.selectbox("Column:", df.columns.tolist(), key="del_col")
            with del_col2:
                del_condition = st.selectbox("Condition:", ["equals", "contains", "greater than", "less than", "empty"], key="del_cond")
            with del_col3:
                del_value = st.text_input("Value:", key="del_val")
            
            if st.button("Preview Deletion", key="preview_del"):
                filtered_df, deleted_count = delete_rows_by_condition(df, del_column, del_condition, del_value)
                st.warning(f"⚠️ This will delete {deleted_count} rows")
                st.write("**Remaining data preview:**")
                show_dataframe_preview(filtered_df)
                
                if st.button("Confirm Deletion", key="confirm_del"):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = del_sheet
                    for r in dataframe_to_rows(filtered_df, index=False, header=True):
                        ws.append(r)
                    
                    st.success(f"✅ Deleted {deleted_count} rows")
                    st.download_button(
                        label="📥 Download Updated File",
                        data=create_download_link(wb, "deleted.xlsx"),
                        file_name="rows_deleted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # Find and Replace
    with st.expander("🔄 Find and Replace"):
        wb = st.session_state[SESSION_WORKBOOK]
        
        find_col1, find_col2 = st.columns(2)
        with find_col1:
            find_text = st.text_input("Find:", key="find_text")
        with find_col2:
            replace_text = st.text_input("Replace with:", key="replace_text")
        
        opt_col1, opt_col2, opt_col3 = st.columns(3)
        with opt_col1:
            match_case = st.checkbox("Match case", key="match_case")
        with opt_col2:
            match_entire = st.checkbox("Match entire cell", key="match_entire")
        with opt_col3:
            search_sheet = st.selectbox("Search in:", ["All sheets"] + wb.sheetnames, key="search_sheet")
        
        if st.button("Preview Replacements", key="preview_replace"):
            if find_text:
                sheet_name = None if search_sheet == "All sheets" else search_sheet
                wb_copy = load_excel_with_password(create_download_link(wb, "temp.xlsx"))
                from openpyxl import load_workbook
                wb_copy = load_workbook(BytesIO(create_download_link(wb, "temp.xlsx")))
                _, replacements_df = find_and_replace(wb_copy, find_text, replace_text, match_case, match_entire, sheet_name)
                
                if not replacements_df.empty:
                    st.info(f"Found {len(replacements_df)} matches")
                    st.dataframe(replacements_df.head(50))
                    
                    if st.button("Confirm Replace", key="confirm_replace"):
                        wb, _ = find_and_replace(wb, find_text, replace_text, match_case, match_entire, sheet_name)
                        st.session_state[SESSION_WORKBOOK] = wb
                        st.success(f"✅ Replaced {len(replacements_df)} occurrences")
                        st.download_button(
                            label="📥 Download Updated File",
                            data=create_download_link(wb, "replaced.xlsx"),
                            file_name="find_replace.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.info("No matches found")
            else:
                st.warning("Please enter text to find")
