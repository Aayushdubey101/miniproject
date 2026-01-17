"""
Tab 2: Data Analysis & Visualization UI
Charts, statistics, pivot tables, filtering, and search
"""

import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from src.features.data_analysis import (
    create_chart, calculate_statistics, create_pivot_table,
    filter_data, search_in_excel
)
from src.utils.file_handlers import create_download_link
from src.ui.components import show_dataframe_preview
from src.config.settings import SESSION_WORKBOOK, SESSION_DF_DICT


def render_data_analysis_tab():
    """Render the Data Analysis & Visualization tab"""
    st.header("📊 Data Analysis & Visualization")
    
    if not st.session_state.get(SESSION_DF_DICT):
        st.warning("⚠️ Please upload an Excel file in the Basic Operations tab first")
        return
    
    # Chart Generation
    with st.expander("📈 Chart Generation", expanded=True):
        chart_sheet = st.selectbox("Select sheet:", list(st.session_state[SESSION_DF_DICT].keys()), key="chart_sheet")
        df = st.session_state[SESSION_DF_DICT][chart_sheet]
        
        chart_col1, chart_col2, chart_col3 = st.columns(3)
        with chart_col1:
            chart_type = st.selectbox("Chart type:", ["Bar Chart", "Line Chart", "Pie Chart", "Scatter Plot"], key="chart_type")
        with chart_col2:
            x_column = st.selectbox("X-axis / Names:", df.columns.tolist(), key="x_col")
        with chart_col3:
            y_column = st.selectbox("Y-axis / Values:", df.columns.tolist(), key="y_col")
        
        chart_title = st.text_input("Chart title:", value=f"{chart_type} - {y_column} by {x_column}", key="chart_title")
        
        if st.button("Generate Chart", key="gen_chart"):
            fig = create_chart(df, chart_type, x_column, y_column, chart_title)
            if fig:
                st.plotly_chart(fig, use_container_width=True)
    
    # Statistics
    with st.expander("📊 Statistical Calculations"):
        stats_sheet = st.selectbox("Select sheet:", list(st.session_state[SESSION_DF_DICT].keys()), key="stats_sheet")
        df = st.session_state[SESSION_DF_DICT][stats_sheet]
        
        selected_columns = st.multiselect("Select columns:", df.columns.tolist(), key="stats_cols")
        
        if st.button("Calculate Statistics", key="calc_stats"):
            if selected_columns:
                stats_df = calculate_statistics(df, selected_columns)
                if stats_df is not None:
                    st.dataframe(stats_df, use_container_width=True)
                    
                    if st.button("Save Statistics to New Sheet", key="save_stats"):
                        wb = st.session_state[SESSION_WORKBOOK]
                        new_sheet = wb.create_sheet(title="Statistics")
                        for r in dataframe_to_rows(stats_df, index=True, header=True):
                            new_sheet.append(r)
                        
                        st.download_button(
                            label="📥 Download with Statistics",
                            data=create_download_link(wb, "statistics.xlsx"),
                            file_name="excel_with_statistics.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.warning("Please select at least one column")
    
    # Pivot Table
    with st.expander("🔄 Pivot Table Creator"):
        pivot_sheet = st.selectbox("Select sheet:", list(st.session_state[SESSION_DF_DICT].keys()), key="pivot_sheet")
        df = st.session_state[SESSION_DF_DICT][pivot_sheet]
        
        piv_col1, piv_col2, piv_col3, piv_col4 = st.columns(4)
        with piv_col1:
            index_col = st.selectbox("Rows (Index):", df.columns.tolist(), key="pivot_index")
        with piv_col2:
            columns_col = st.selectbox("Columns:", df.columns.tolist(), key="pivot_cols")
        with piv_col3:
            values_col = st.selectbox("Values:", df.columns.tolist(), key="pivot_vals")
        with piv_col4:
            aggfunc = st.selectbox("Aggregation:", ["sum", "mean", "count", "min", "max"], key="pivot_agg")
        
        if st.button("Create Pivot Table", key="create_pivot"):
            pivot_df = create_pivot_table(df, index_col, columns_col, values_col, aggfunc)
            if pivot_df is not None:
                st.dataframe(pivot_df, use_container_width=True)
                
                if st.button("Save Pivot to New Sheet", key="save_pivot"):
                    wb = st.session_state[SESSION_WORKBOOK]
                    new_sheet = wb.create_sheet(title="Pivot_Table")
                    for r in dataframe_to_rows(pivot_df, index=True, header=True):
                        new_sheet.append(r)
                    
                    st.download_button(
                        label="📥 Download with Pivot Table",
                        data=create_download_link(wb, "pivot.xlsx"),
                        file_name="excel_with_pivot.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # Filter & Sort
    with st.expander("🔍 Filter & Sort Data"):
        filter_sheet = st.selectbox("Select sheet:", list(st.session_state[SESSION_DF_DICT].keys()), key="filter_sheet")
        df = st.session_state[SESSION_DF_DICT][filter_sheet]
        
        st.write("**Filter Options**")
        filt_col1, filt_col2, filt_col3 = st.columns(3)
        with filt_col1:
            filter_column = st.selectbox("Column:", df.columns.tolist(), key="filter_col")
        with filt_col2:
            filter_condition = st.selectbox("Condition:", ["equals", "contains", "greater than", "less than", "not equals"], key="filter_cond")
        with filt_col3:
            filter_value = st.text_input("Value:", key="filter_val")
        
        st.write("**Sort Options**")
        sort_col1, sort_col2 = st.columns(2)
        with sort_col1:
            sort_column = st.selectbox("Sort by:", df.columns.tolist(), key="sort_col")
        with sort_col2:
            sort_order = st.radio("Order:", ["Ascending", "Descending"], key="sort_order", horizontal=True)
        
        if st.button("Apply Filter & Sort", key="apply_filter"):
            result_df = df.copy()
            if filter_value:
                result_df = filter_data(result_df, filter_column, filter_condition, filter_value)
            result_df = result_df.sort_values(by=sort_column, ascending=(sort_order == "Ascending"))
            
            st.success(f"Filtered to {len(result_df)} rows from {len(df)} total rows")
            show_dataframe_preview(result_df)
            
            if st.button("Save Filtered Data", key="save_filtered"):
                wb = Workbook()
                ws = wb.active
                ws.title = "Filtered_Data"
                for r in dataframe_to_rows(result_df, index=False, header=True):
                    ws.append(r)
                
                st.download_button(
                    label="📥 Download Filtered Data",
                    data=create_download_link(wb, "filtered.xlsx"),
                    file_name="filtered_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    # Search
    with st.expander("🔎 Search Functionality"):
        search_term = st.text_input("Search for:", key="search_term")
        case_sensitive = st.checkbox("Case sensitive", key="case_sens")
        
        if st.button("Search", key="search_btn"):
            if search_term and st.session_state.get(SESSION_WORKBOOK):
                with st.spinner("Searching..."):
                    results_df = search_in_excel(st.session_state[SESSION_WORKBOOK], search_term, case_sensitive)
                
                if not results_df.empty:
                    st.success(f"Found {len(results_df)} matches")
                    st.dataframe(results_df, use_container_width=True)
                else:
                    st.info("No matches found")
            else:
                st.warning("Please enter a search term")
